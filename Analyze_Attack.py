
import os
import xlwt
import pandas as pd
import openpyxl as op
from xlwt import Workbook


from WebRequest import web_request
from WriteTextFile import write_text_file
from RegularExpression import regular_expression
from FindContexts import find_contexts
from GenerateFormUrls import generate_form_urls_with_payloads
from ContextEncoding import context_encoding
from AttackMethodology import attack_methodology

class analyze_attack:
    payload = 'abc/uvw"' + "xyz'yxz<zxy"
    Text = None 
    folder = base = ''

    def __init__(self, base, folder):
        self.base = base
        self.folder = folder
        self.create_text_file()

    def create_text_file(self):
        path = self.base + '/AaallAttacks' + '.txt'
        textfile = open(path, "w")
        textfile.write(str(self.base)+'\n\n')
        textfile.close()

    def get_source(self,url):
        Web = web_request(url,'GET')
        source = Web.get_source()
        return source
 
    def read_excel(self):
        df = pd.read_excel(self.folder + '/file.xlsx')
        links = df['Attack URL']
        return links

    def write_excel_attack_description(self, attack_url, context, status, detection):
        self.write_attacks_text_file(attack_url, context, status, detection)
        return 
        wb = op.load_workbook(self.folder + '/file.xlsx')
        ws = wb['Sheet1']

        if not detection == 'None':
            ws.append([str(attack_url), context, status, ''])
            for data in detection:
                new = ws.cell(row=ws.max_row, column=4).value + str(data) + ' , '
                ws.cell(row=ws.max_row, column=4).value = new
                pass 
        else:
            ws.append([str(attack_url), context, status, str(detection)])
        
        wb.save(self.folder + '/file.xlsx')
        wb.close()
    
    def write_attacks_text_file(self,attack_url, context, Successful, detection):
        breakline = '\n-----------------------------------------------------------------------------------------------------\n'
        path = self.base + '/AaallAttacks' + '.txt'
        textfile = open(path, "a")
        if not detection == 'None':
            textfile.write(str(attack_url) + '\nContext = ' + context + '\nSuccess = ' + str(Successful) + '\nDetection:\n')
            for d in detection: textfile.write(str(d.encode(encoding='UTF-8')) + '\n')
        else: 
            textfile.write(str(attack_url) + '\nContext = ' + context + '\nSuccess = ' + str(Successful)  + '\nDetection:\n' + 'None')

        textfile.write(breakline)

    def remove_duplicate_get_urls(self,get_urls):
        unique_get_urls = [] 
        # print('Removing Duplicates...')
        for x in get_urls:
            # Removing Duplicates
            if x not in unique_get_urls: unique_get_urls.append(x)  
        return unique_get_urls

    def collect_data(self,links):
        Search = generate_form_urls_with_payloads()
        get_urls = []    # it is the list in which GET <form> links are added / concatenated
        links_count = 0
        
        # we have 2 links
        for link in links:  #links contains all the references/links
            links_count += 1
            print( '\n Webpage [', links_count , '/',  len(links), '] => ', link, end="")
            if not link == [] and link.__contains__('http'):   
                # generates links/urls from every <form ... method="get">. 
                # We receive a list of links/urls. 
                get_params, get_urls = Search.start_search(link) 
                self.get_params = get_params
            else: continue
            unique_get_urls = self.remove_duplicate_get_urls(get_urls)
            print('\nUnique GET URLs:[',len(unique_get_urls), ']') #, unique_get_urls)
            self.collect_response_data(link, unique_get_urls)

    def collect_response_data(self,link, unique_get_urls): # adds payload in GET params..?
        attack_urls = []
        count = 0

        # Text File Containing GET Form Response.
        self.Text = write_text_file(self.base,link, self.payload)
        self.Text.write_directly( '\nGET Params:[' + str(len(self.get_params)) + ']\n' + str( self.get_params )  + '\n')
        # print( 'Unique GET URLs [', len(unique_get_urls) , ']\n', unique_get_urls ,'\n' )

        # Collecting Context Data of  <GET Forms>
        for u in unique_get_urls:
            count+=1
            if u: url = u   # if url is not Empty then Attach the Payload
            else: continue
            if url.__contains__('http'):    # Checking for a Valid Url
                print('[', count , '/', len(unique_get_urls), ']', 'Analysing GET URL =>',  url )
                source = self.get_source(url) # Retrieving source code of the webpage
                # Fucntion Call
                self.record_data(url,source)

    # Finding Contexts, selecting attack-Methodology and attack-payloads
    def record_data(self,url,source):
        Find = find_contexts()
        # Finding All Contexts
        attrs, htmls, scripts, urls, same_attrs, same_htmls, same_scripts, same_urls = Find.find_context(url, self.payload, str(source) )
        # Writing Contexts to Text File
        self.Text.write_contexts(url, attrs, htmls, scripts, urls, same_attrs, same_htmls, same_scripts, same_urls)
        # Cotext Encoding and Attack Methodology for each Context
        for attr in attrs       :   self.check_encoding_and_attack( url, 'ATTR', attr)
        for html in htmls       :   self.check_encoding_and_attack( url, 'HTML', html)
        for script in scripts   :   self.check_encoding_and_attack( url, 'SCRIPT', script)
        for my_url in urls      :   self.check_encoding_and_attack( url, 'URL', my_url)

    def check_encoding_and_attack(self, url, context_name, context_data):
        CE = context_encoding(self.Text)
        # For each Context { encoding_analyzer() } returns True for encoding or escaping of special chars and False otherwise.
        presence, double_quotes, single_quotes, lessthan_sign, forwardslash = CE.encoding_analyzer(context_name, context_data)
        if context_name == 'URL' or context_name == 'HTML' or context_name == 'ATTR':
            print_presence = str(presence) + '  '
        else:   
            print_presence = str(presence)
        
        # print('Special Chars =   \t\"\t \'\t<\t(')
        # print(context_name,'Mititgation\t', double_quotes, single_quotes, lessthan_sign, forwardslash )
        
        self.Text.write_encoding(context_name, presence, double_quotes, single_quotes, lessthan_sign, forwardslash)
        self.try_attacks(url, context_name, presence, double_quotes, single_quotes, lessthan_sign, forwardslash)

    def try_attacks(self, url, context_name, presence, double_quotes, single_quotes, lessthan_sign, forwardslash ):
        AM = attack_methodology()
        pay = self.payload
        attack_payloads = []
        tag, attack_payloads = AM.get_attack_payload(context_name, presence, double_quotes, single_quotes, lessthan_sign, forwardslash )
        # print('Attack Payloads: ', attack_payloads)
        
        if tag:
            if attack_payloads:
                self.Text.write_directly('\nAttack Payloads for ' + str(context_name) + '\n' + str(attack_payloads) + '\n')
                print('\nAttack Paloads for ', context_name, '\n', attack_payloads , '\n')
            for attack in attack_payloads:
                url = url.replace(pay, attack)
                pay = attack

                links = self.read_excel()
                leave = False
                for link in links:
                    # if url == link: leave = True
                    pass
                if leave:   
                    print('\n\t Duplicate = ', url, '\n')
                    continue

                print('\n', context_name, 'Attack Url: ', url)
                self.Text.write_directly('\n'+ context_name + ' Attack Url: ' + url)
                data = self.get_source(url)
                if( str(data).__contains__(attack)):
                    print('\n\n=>Detection  Successful with Payload: ', str(attack), '\n')
                    self.Text.write_directly('\n\n=>Detection  Successful with Payload: ' + str(attack) + '\n')
                    # print('=>The Automated Tool Assumes that there is a potential XSS Present in the Website\n')
                    RegExp = regular_expression(data)
                    RegExp.set_payload(attack)
                    value = RegExp.context_attack(context_name)

                    CE = context_encoding(self.Text)
                    detection = []
                    print('FINAL OUTPUT:\n')
                    for val in value:
                        if  context_name == 'ATTR':
                            if not single_quotes and not CE.attr_double_quotes_outside(val, attack): detection.append(str(val))
                            if not double_quotes and not CE.attr_single_quotes_outside(val, attack): detection.append(str(val))
                        elif  context_name == 'SCRIPT':
                            if not single_quotes and not CE.script_double_quotes_outside(val, attack): detection.append(str(val))
                            if not double_quotes and not CE.script_single_quotes_outside(val, attack): detection.append(str(val))
                        else:
                            detection.append(str(val))
                     
                    self.write_excel_attack_description(url, context_name, 'TRUE', detection)
                    self.Text.write_directly('\nFINAL OUTPUT: ' + '\n')
                    print( detection  , '\n\n')
                    for d in detection: 
                        self.Text.write_directly(str(d) + "\n")
                else:
                    print('\n\n ______Detection  UnSuccessful with payload: ', attack, '\n\n')
                    self.write_excel_attack_description(url, context_name, 'FALSE', 'None')
                    self.Text.write_directly('\n\n ______Detection  UnSuccessful with payload: ' + str(attack) + '\n\n')
        
  
